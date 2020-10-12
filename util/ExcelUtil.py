from openpyxl import Workbook, load_workbook
import random
import string


def read_excel(fname):
    wb = load_workbook(filename=fname)
    data = wb.active
    r = data.max_row
    col = data.max_column
    item = []
    array = []
    for i in range(1, r + 1):
        for j in range(2, col + 1):
            cell_value = data.cell(row=i, column=j)
            item.append(cell_value.value)
        item.append(generate_pin_password())
        array.append(tuple(item))
        item = []
    print(array)
    return array


def write_to_excel(array):
    wb = Workbook()
    sheet = wb.active

    for item in array:
        sheet.append(item)
    print("successfully inserted to excel file gen.xlsx")
    wb.save("gen.xlsx")


def generate_pin_excel(count, filename):
    wb = Workbook()
    sheet = wb.active
    item = []
    array = []
    for i in range(1, count+1):
        item.append(generate_pin_password() + "-" + generate_pin_password())
        sheet.append(tuple(item))
        array.append(tuple([item[0], i]))
        item = []
    wb.save(filename + '.xlsx')
    return array


def generate_pin_password():
    alphanum = string.ascii_uppercase + string.digits
    pin = ""
    for i in range(4):
        randchar = random.randrange(0, len(alphanum))
        pin += alphanum[randchar]
    return pin
